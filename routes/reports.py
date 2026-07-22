import io
from flask import Blueprint, request, send_file, jsonify
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from db import get_db
from routes.auth import jwt_required

reports_bp = Blueprint('reports', __name__, url_prefix='/api/admin/export')

@reports_bp.route('/excel', methods=['GET'])
@jwt_required(role='admin')
def export_excel():
    month = request.args.get('month', datetime.now().strftime('%Y-%m')).strip()
    department = request.args.get('department', '').strip()

    conn = get_db()
    cursor = conn.cursor()

    query = """
        SELECT a.date, u.employee_id, u.name as user_name, u.department, u.designation,
               COALESCE(a.check_in, 'N/A') as check_in,
               COALESCE(a.check_out, 'N/A') as check_out,
               a.total_hours, a.break_duration_mins, a.net_hours, a.status,
               CASE WHEN a.is_late = 1 THEN 'Yes' ELSE 'No' END as is_late
        FROM attendance a
        JOIN users u ON a.user_id = u.id
        WHERE a.date LIKE ?
    """
    params = [f"{month}%"]

    if department:
        query += " AND u.department = ?"
        params.append(department)

    query += " ORDER BY a.date ASC, u.employee_id ASC"

    cursor.execute(query, params)
    records = cursor.fetchall()
    conn.close()

    # Create Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance_{month}"

    # Header styling
    title_font = Font(name='Segoe UI', size=16, bold=True, color='1F2937')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Title Block
    ws.merge_cells('A1:L1')
    ws['A1'] = f"Office Attendance Report - {month}"
    ws['A1'].font = title_font
    ws['A1'].alignment = left_align
    ws.row_dimensions[1].height = 35

    # Column Headers
    headers = ['Date', 'Emp ID', 'Employee Name', 'Department', 'Designation', 'Check In', 'Check Out', 'Total Hrs', 'Break (Mins)', 'Net Hrs', 'Status', 'Late?']
    ws.append([]) # Row 2 blank
    ws.append(headers) # Row 3

    ws.row_dimensions[3].height = 25

    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Data Rows
    for row_idx, r in enumerate(records, start=4):
        ws.append([
            r['date'], r['employee_id'], r['user_name'], r['department'], r['designation'],
            r['check_in'], r['check_out'], r['total_hours'], r['break_duration_mins'],
            r['net_hours'], r['status'], r['is_late']
        ])
        ws.row_dimensions[row_idx].height = 20
        
        # Format row cells
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = thin_border
            if c in [1, 2, 6, 7, 8, 9, 10, 11, 12]:
                cell.alignment = center_align

    # Column widths auto-adjust
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Attendance_Report_{month}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@reports_bp.route('/pdf', methods=['GET'])
@jwt_required(role='admin')
def export_pdf():
    month = request.args.get('month', datetime.now().strftime('%Y-%m')).strip()
    department = request.args.get('department', '').strip()

    conn = get_db()
    cursor = conn.cursor()

    query = """
        SELECT a.date, u.employee_id, u.name as user_name, u.department,
               COALESCE(a.check_in, 'N/A') as check_in,
               COALESCE(a.check_out, 'N/A') as check_out,
               a.net_hours, a.status
        FROM attendance a
        JOIN users u ON a.user_id = u.id
        WHERE a.date LIKE ?
    """
    params = [f"{month}%"]

    if department:
        query += " AND u.department = ?"
        params.append(department)

    query += " ORDER BY a.date ASC, u.employee_id ASC LIMIT 150"

    cursor.execute(query, params)
    records = cursor.fetchall()
    conn.close()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1F2937'),
        spaceAfter=12
    )

    story.append(Paragraph(f"Office Attendance Report ({month})", title_style))
    if department:
        story.append(Paragraph(f"Department: {department}", styles['Normal']))
    story.append(Spacer(1, 15))

    # Table data
    data = [['Date', 'Emp ID', 'Name', 'Department', 'Check In', 'Check Out', 'Net Hrs', 'Status']]
    for r in records:
        data.append([
            r['date'],
            r['employee_id'],
            r['user_name'],
            r['department'],
            r['check_in'].split(' ')[1] if ' ' in r['check_in'] else r['check_in'],
            r['check_out'].split(' ')[1] if ' ' in r['check_out'] else r['check_out'],
            f"{r['net_hours']}h",
            r['status']
        ])

    table = Table(data, colWidths=[65, 55, 100, 85, 60, 60, 50, 55])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9FAFB')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    story.append(table)
    doc.build(story)

    buffer.seek(0)
    filename = f"Attendance_Summary_{month}.pdf"
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )
